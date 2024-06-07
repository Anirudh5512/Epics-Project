import tkinter as tk
from tkinter import *
from tkinter import messagebox

import openpyxl
from fuzzywuzzy import fuzz
import os
print("Current working directory:", os.getcwd())


ans = 0
text = ''
kt = 0
cm = 0
gm = 0
fr = 0
g = 0
file = ""
fileQ = ""
Qtext = ""
frf = 1
ktf = 0
cmf = 0
gmf = 0
strans = ""
ansl = []
keyword = []
com = ["and", "that", "the", "for", "it", "it's", "was", "his", "who", "work", "used", "way", "also", "by", "can",
       "which", "as", "known", "then", "if", "between", "through", "another", "", "or", "my", "in", "from", "a", "any",
       "on", "combination", "to", "into", "is", "of", "It", "A", "each", "both"]


def load_words():
    with open('words_alpha.json') as word_file:
        valid_words = set(word_file.read().split())
    return valid_words


def ans_key(s):
    global strans
    strans = ""
    global ansl
    ansl.clear()
    global keyword
    keyword.clear()
    for i in range(6, s.max_row + 1):
        t = s.cell(row=i, column=1).value
        ansl.append(t)
        strans = strans + " \n\n" + str(i - 5) + ")" + t

    for a in ansl:
        ass = a.split()
        for sas in ass:
            sas = sas.lower()
            for check in ansl:
                if a == check:
                    continue
                else:
                    assc = check.split()
                    for x in assc:
                        x = x.lower()
                        if x == sas and x not in com and x not in keyword:
                            keyword.append(x)


class Test(tk.Frame):
    def __init__(self):
        new = tk.Frame.__init__(self)
        new = Toplevel(self)

        new.title("Question Paper")

        self.tk_setPalette(background='#ececec')

        new.geometry("610x377+361+223")

        global Qtext
        tk.Label(new, text=Qtext).grid(row=0, column=0, padx=13, pady=21, sticky=W)

        self.entryA = Text(new, height=14, width=46, padx=5, pady=5, wrap=WORD, background='white')
        self.entryA.grid(row=0, column=1, padx=13, pady=21, sticky=E)

        tk.Button(new, text='  Submit  ', default='active', command=self.click_ok).grid(row=1, column=1, padx=13, pady=21)

    def click_ok(self):
        global text
        text = self.entryA.get("1.0", END)

        if text == "" or text == " ":
            messagebox.showinfo("Blank Input Error", "Please enter a Statement")
        elif len(text.split()) < 5:
            messagebox.showinfo("Too Short Input Error", "Please enter a proper Statement")

        else:
            self.newWindow = Report()


class App(tk.Frame):
    def __init__(self, master):
        tk.Frame.__init__(self, master)
        self.pack()
        self.master.resizable(False, False)
        self.master.tk_setPalette(background='#ececec')

        self.master.geometry("500x190+500+323")

        self.master.title("Automatic Answer checker")

        tkt = StringVar(root)

        choices = {'E-Commerce', 'NLP', 'Cryptography', 'Cyber-Security', 'Philosophy'}
        tkt.set('E-Commerce')
        openmyfile('E-Commerce')
        popupMenu = OptionMenu(self, tkt, *choices)
        Label(self, text="Subject     : ").grid(row=2, column=0, padx=5, pady=10, sticky=W)
        popupMenu.grid(row=2, column=1)
        global file
        global loc
        global Qtext
        global frf
        global ktf
        global cmf
        global gmf

        def change_dropdown(*args):
            file = tkt.get()
            openmyfile(file)

        tkt.trace('w', change_dropdown)

        tk.Label(self, text="Username : ").grid(row=0, column=0, padx=5, pady=10, sticky=W)

        self.entryA = tk.Entry(self, width=26, background='white')
        self.entryA.grid(row=0, column=1, padx=5, pady=10, sticky=W)
        self.entryA.focus_set()
        tk.Label(self, text="Password : ").grid(row=1, column=0, padx=5, pady=10, sticky=W)

        self.entryB = tk.Entry(self, width=26, background='white', show="*")
        self.entryB.grid(row=1, column=1, padx=5, pady=15, sticky=W)

        tk.Button(self, text='  Submit  ', default='active', command=self.click_ok).grid(row=3, column=1, padx=13, pady=21,
                                                                                        sticky=S)

    def click_ok(self):
        user = self.entryA.get()
        password = self.entryB.get()

        op = ("login.xlsx")
        wbr = openpyxl.load_workbook(op, read_only=True)
        sh = wbr.active
        cuser = sh[1]
        cpass = sh[3]

        if user in cuser and password in cpass:
            self.newWindow = Test()
            root.withdraw()
        else:
            messagebox.showinfo("Login error", "Please enter the correct credentials")


class Report(tk.Frame):
    def __init__(self):
        new = tk.Frame.__init__(self)
        new = Toplevel(self)

        new.geometry("350x180+500+300")
        new.title("Evaluation Report")
        global ans
        ans = 0
        global ansl
        global text
        global keyword
        g = 0
        text = text.strip()
        english_words = load_words()
        if text == " " or text == "":
            gm = 0
        else:
            text = text.split()
            for t in text:
                t = t.lower()
                if t[-1] == ".":
                    t = t[:-1]
                if t in english_words:
                    g = g + 1

        if g > 7:
            for ev in ansl:
                ans = ans + fuzz.token_set_ratio(ev, text)
                ans = ans + fuzz.ratio(ev, text)

        global kt
        global cm
        repeat = []
        for t in text:
            t = t.lower()
            if t in keyword and t not in repeat:
                value = keyword.index(t)
                if value >= 2:
                    kt = kt + 0.05
                elif value == 0:
                    kt = kt + 0.1
                elif value == 1:
                    kt = kt + 0.08
                repeat.append(t)

        check = []
        c = 0
        for i in range(0, len(repeat) - 1):
            if keyword.index(repeat[i]) < keyword.index(repeat[i + 1]) and repeat[i] not in check:
                check.append(repeat[i])
                check.append(repeat[i + 1])

        if kt > 1:
            kt = 1
        if len(check) == 0:
            cm = 0
        else:
            cm = len(check) / len(repeat)

        global fr
        global x
        gm = g / len(text)
        fr = ans / (len(ansl))
        global frf, ktf, cmf, gmf

        r = fr / (frf * 100) + ktf * kt + cmf * cm + gmf * gm

        if r > 0.95:
            r = 10
        elif r > 0.9:
            r = 9.5
        elif r > 0.85:
            r = 9
        elif r > 0.8:
            r = 8.5
        elif r > 0.75:
            r = 8
        elif r > 0.7:
            r = 7.5
        elif r > 0.65:
            r = 7
        elif r > 0.6:
            r = 6.5
        elif r > 0.55:
            r = 6
        elif r > 0.5:
            r = 5.5
        elif r > 0.45:
            r = 5
        elif r > 0.4:
            r = 4.5
        elif r > 0.35:
            r = 4
        elif r > 0.3:
            r = 3.5
        elif r > 0.25:
            r = 3
        elif r > 0.2:
            r = 2.5
        elif r > 0.15:
            r = 2
        elif r > 0.1:
            r = 1.5
        elif r > 0.05:
            r = 1
        else:
            r = 0

        ans = (str)(r)
        new.label = tk.Label(new, text=" Your Marks is = " + ans)
        new.label.grid(row=2, column=2, padx=10, pady=15, sticky=N)
        new.button = tk.Button(new, text="Detailed Report", width=15, command=self.Det_report)
        new.button.grid(row=3, column=3, padx=10, pady=15, sticky=N)
        new.button2 = tk.Button(new, text="Close", width=15, command=self.close_window)
        new.button2.grid(row=4, column=3, padx=10, pady=15, sticky=N)

    def Det_report(self):
        self.newWindow = Det_Report()

    def close_window(self):
        self.master.destroy()


class Det_Report(tk.Frame):
    def __init__(self):
        new = tk.Frame.__init__(self)
        new = Toplevel(self)

        new.geometry("700x650+361+100")
        new.title("Evaluation Full Report")
        global ans
        ans = 0
        global ansl
        global text
        global keyword
        global strans
        g = 0
        global loc
        global fr
        text = text.strip()
        english_words = load_words()
        tk.Label(new, text="Marks Evaluation Report", font=("Helvetica", 16), pady=20).grid(row=0, column=1,
                                                                                            columnspan=2)
        if text == " " or text == "":
            tk.Label(new, text="Plagiarised Content : \nNo\n\nGrammar Mistakes : \nNo\n\n",
                     font=("Helvetica", 12)).grid(row=1, column=1, padx=10, pady=10)
        else:
            tk.Label(new, text="Your Content is : \n\n" + text, font=("Helvetica", 12)).grid(row=1, column=1,
                                                                                             padx=10, pady=10, sticky=W)
            text = text.split()
            for t in text:
                t = t.lower()
                if t[-1] == ".":
                    t = t[:-1]
                if t in english_words:
                    g = g + 1
            if g == 0:
                tk.Label(new, text="\n\nPlagiarised Content : \nYes\n\n", font=("Helvetica", 12)).grid(row=2,
                                                                                                      column=1, padx=10,
                                                                                                      pady=10,
                                                                                                      sticky=W)
            else:
                tk.Label(new, text="\n\nPlagiarised Content : \nNo\n\n", font=("Helvetica", 12)).grid(row=2,
                                                                                                     column=1, padx=10,
                                                                                                     pady=10,
                                                                                                     sticky=W)

        if g > 7:
            for ev in ansl:
                ans = ans + fuzz.token_set_ratio(ev, text)
                ans = ans + fuzz.ratio(ev, text)

        global kt
        global cm
        repeat = []
        for t in text:
            t = t.lower()
            if t in keyword and t not in repeat:
                value = keyword.index(t)
                if value >= 2:
                    kt = kt + 0.05
                elif value == 0:
                    kt = kt + 0.1
                elif value == 1:
                    kt = kt + 0.08
                repeat.append(t)

        check = []
        c = 0
        for i in range(0, len(repeat) - 1):
            if keyword.index(repeat[i]) < keyword.index(repeat[i + 1]) and repeat[i] not in check:
                check.append(repeat[i])
                check.append(repeat[i + 1])

        if kt > 1:
            kt = 1
        if len(check) == 0:
            cm = 0
        else:
            cm = len(check) / len(repeat)

        global fr
        global x
        gm = g / len(text)
        fr = ans / (len(ansl))
        global frf, ktf, cmf, gmf

        r = fr / (frf * 100) + ktf * kt + cmf * cm + gmf * gm

        tk.Label(new, text="Grammar Mistakes : \n\nTotal English Words : " + str(g) + "\n\nMistakes  : "
                           + str(len(text) - g) + "\n\n", font=("Helvetica", 12)).grid(row=4, column=1, padx=10, pady=10,
                                                                                        sticky=W)

        if g > 7:
            tk.Label(new, text="\n\nPlagiarism Matches : " + str(g) + " out of " + str(len(text))
                               + "\n\n", font=("Helvetica", 12)).grid(row=5, column=1, padx=10, pady=10, sticky=W)
        else:
            tk.Label(new, text="\n\nPlagiarism Matches : 0 out of " + str(len(text))
                               + "\n\n", font=("Helvetica", 12)).grid(row=5, column=1, padx=10, pady=10, sticky=W)

        tk.Label(new, text="\n\nSubject Specific Keywords : " + str(len(check)) + " out of " + str(
            len(repeat)) + "\n\n", font=("Helvetica", 12)).grid(row=6, column=1, padx=10, pady=10, sticky=W)

        tk.Label(new, text="\n\nGrammar Marks : " + str(gm) + " out of 1\n\n", font=("Helvetica", 12)).grid(row=7, column=1,
                                                                                                          padx=10,
                                                                                                          pady=10,
                                                                                                          sticky=W)

        tk.Label(new, text="\n\nKeyword Marks : " + str(kt) + " out of 1\n\n", font=("Helvetica", 12)).grid(row=8, column=1,
                                                                                                           padx=10,
                                                                                                           pady=10,
                                                                                                           sticky=W)

        tk.Label(new, text="\n\nCommon Word Marks : " + str(cm) + " out of 1\n\n", font=("Helvetica", 12)).grid(row=9,
                                                                                                                column=1,
                                                                                                                padx=10,
                                                                                                                pady=10,
                                                                                                                sticky=W)

        tk.Label(new, text="\n\nFinal Score : " + str(r * 10) + " out of 10\n\n",
                 font=("Helvetica", 12)).grid(row=10, column=1, padx=10, pady=10, sticky=W)

        tk.Label(new, text="\n\n" + strans, font=("Helvetica", 12)).grid(row=2, column=2, rowspan=8, padx=10, pady=10,
                                                                        sticky=W)

        new.button2 = tk.Button(new, text="Close", width=15, command=self.close_window)
        new.button2.grid(row=11, column=2, padx=10, pady=15, sticky=N)

    def close_window(self):
        self.master.destroy()


def openmyfile(sub):
    global loc
    if sub == 'E-Commerce':
        loc = ("E-Commerce.xlsx")
        x = 1

    elif sub == 'NLP':
        loc = ("NLP.xlsx")
        x = 2

    elif sub == 'Cryptography':
        loc = ("Cryptography.xlsx")
        x = 3

    elif sub == 'Cyber-Security':
        loc = ("Cyber-Security.xlsx")
        x = 4

    elif sub == 'Philosophy':
        loc = ("Philosophy.xlsx")
        x = 5

    global wb
    global sheet
    wb = openpyxl.load_workbook(loc)
    sheet = wb.active


root = tk.Tk()
app = App(master=root)
app.mainloop()
